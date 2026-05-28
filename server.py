from __future__ import annotations

import json
import os
import re
import shutil
import urllib.parse
import urllib.request
from collections import Counter
from datetime import datetime
from pathlib import Path
from typing import Any

from mcp.server.fastmcp import FastMCP
from pydantic import BaseModel


APP_NAME = "Document Organizer MCP"
DEFAULT_ALLOWED_EXTENSIONS = {
    ".txt",
    ".md",
    ".json",
    ".csv",
    ".pdf",
    ".docx",
    ".xlsx",
}

mcp = FastMCP(APP_NAME, json_response=True)


class FileMatch(BaseModel):
    path: str
    name: str
    extension: str
    size_bytes: int


class SearchMatch(BaseModel):
    path: str
    occurrences: int
    snippet: str


class RenameSuggestion(BaseModel):
    current_path: str
    suggested_name: str
    suggested_path: str
    basis: str


def _doc_root() -> Path:
    raw_root = os.environ.get("DOC_ROOT", "./data")
    return Path(raw_root).expanduser().resolve()


def _public_base_url() -> str:
    return os.environ.get("PUBLIC_BASE_URL", "").rstrip("/")


def _allowed_extensions() -> set[str]:
    raw = os.environ.get("DOC_EXTENSIONS", "")
    if not raw.strip():
        return DEFAULT_ALLOWED_EXTENSIONS
    values = {part.strip().lower() for part in raw.split(",") if part.strip()}
    normalized = {value if value.startswith(".") else f".{value}" for value in values}
    return normalized or DEFAULT_ALLOWED_EXTENSIONS


def _resolve_user_path(user_path: str) -> Path:
    candidate = (_doc_root() / user_path).resolve()
    try:
        candidate.relative_to(_doc_root())
    except ValueError as exc:
        raise ValueError("Requested path is outside DOC_ROOT.") from exc
    return candidate


def _iter_document_paths(subdir: str = "") -> list[Path]:
    root = _resolve_user_path(subdir) if subdir else _doc_root()
    if not root.exists():
        raise FileNotFoundError(f"Path does not exist: {root}")
    extensions = _allowed_extensions()
    return sorted(
        [
            path
            for path in root.rglob("*")
            if path.is_file() and path.suffix.lower() in extensions
        ]
    )


def _relative_path(path: Path) -> str:
    return str(path.relative_to(_doc_root()))


def _local_document_url(path: Path) -> str:
    relative = _relative_path(path).replace("\\", "/")
    if _public_base_url():
        return f"{_public_base_url()}/documents/{urllib.parse.quote(relative)}"
    return f"doc://local/{relative}"


def _read_text_file(path: Path) -> str:
    return path.read_text(encoding="utf-8", errors="ignore")


def _read_json_file(path: Path) -> str:
    data = json.loads(path.read_text(encoding="utf-8", errors="ignore"))
    return json.dumps(data, indent=2, ensure_ascii=True)


def _read_pdf_file(path: Path) -> str:
    from pypdf import PdfReader

    reader = PdfReader(str(path))
    pages = [(page.extract_text() or "").strip() for page in reader.pages]
    return "\n\n".join(page for page in pages if page)


def _read_docx_file(path: Path) -> str:
    from docx import Document

    document = Document(str(path))
    paragraphs = [paragraph.text.strip() for paragraph in document.paragraphs]
    return "\n".join(line for line in paragraphs if line)


def _read_xlsx_file(path: Path) -> str:
    from openpyxl import load_workbook

    workbook = load_workbook(filename=str(path), data_only=True, read_only=True)
    sections: list[str] = []
    for sheet in workbook.worksheets:
        rows: list[str] = []
        for row in sheet.iter_rows(values_only=True):
            values = ["" if cell is None else str(cell) for cell in row]
            if any(value.strip() for value in values):
                rows.append(" | ".join(values).strip())
        if rows:
            sections.append(f"# Sheet: {sheet.title}\n" + "\n".join(rows))
    return "\n\n".join(sections)


def _extract_text(path: Path) -> str:
    suffix = path.suffix.lower()
    if suffix in {".txt", ".md", ".csv"}:
        return _read_text_file(path)
    if suffix == ".json":
        return _read_json_file(path)
    if suffix == ".pdf":
        return _read_pdf_file(path)
    if suffix == ".docx":
        return _read_docx_file(path)
    if suffix == ".xlsx":
        return _read_xlsx_file(path)
    raise ValueError(f"Unsupported file type: {suffix}")


def _snippet(text: str, query: str, max_chars: int = 220) -> str:
    lowered = text.lower()
    needle = query.lower()
    index = lowered.find(needle)
    if index < 0:
        compact = " ".join(text.split())
        return compact[:max_chars]

    start = max(index - 80, 0)
    end = min(index + len(query) + 120, len(text))
    segment = " ".join(text[start:end].split())
    return segment[:max_chars]


def _safe_slug(value: str, fallback: str = "document") -> str:
    slug = re.sub(r"[^A-Za-z0-9._-]+", "-", value.strip())
    slug = re.sub(r"-{2,}", "-", slug).strip("-._")
    return slug or fallback


def _first_meaningful_line(text: str) -> str:
    for line in text.splitlines():
        cleaned = " ".join(line.split()).strip()
        if len(cleaned) >= 6:
            return cleaned
    return ""


def _extract_date_hint(text: str) -> str | None:
    patterns = [
        r"\b(20\d{2})[-/](0[1-9]|1[0-2])[-/](0[1-9]|[12]\d|3[01])\b",
        r"\b(0[1-9]|1[0-2])[-/](0[1-9]|[12]\d|3[01])[-/](20\d{2})\b",
    ]
    for pattern in patterns:
        match = re.search(pattern, text)
        if not match:
            continue
        parts = match.groups()
        if len(parts[0]) == 4:
            return f"{parts[0]}-{parts[1]}-{parts[2]}"
        return f"{parts[2]}-{parts[0]}-{parts[1]}"
    return None


def _suggest_name_for_document(path: Path, text: str) -> tuple[str, str]:
    title_line = _first_meaningful_line(text)
    stem_guess = _safe_slug(title_line[:80], fallback=path.stem)
    date_hint = _extract_date_hint(text)
    if date_hint:
        stem_guess = f"{date_hint}_{stem_guess}"
        basis = "first meaningful line plus detected date"
    elif title_line:
        basis = "first meaningful line from extracted text"
    else:
        stem_guess = _safe_slug(path.stem, fallback="document")
        basis = "existing filename because no stronger text cue was found"
    return f"{stem_guess}{path.suffix.lower()}", basis


def _drive_access_token() -> str:
    token = os.environ.get("GOOGLE_DRIVE_ACCESS_TOKEN", "").strip()
    if not token:
        raise ValueError(
            "GOOGLE_DRIVE_ACCESS_TOKEN is not set. Add an OAuth access token to enable Drive tools."
        )
    return token


def _drive_request(
    path: str,
    query: dict[str, Any] | None = None,
    accept: str = "application/json",
) -> tuple[bytes, dict[str, Any]]:
    encoded_query = urllib.parse.urlencode(
        {key: value for key, value in (query or {}).items() if value is not None}
    )
    url = f"https://www.googleapis.com{path}"
    if encoded_query:
        url = f"{url}?{encoded_query}"

    request = urllib.request.Request(
        url,
        headers={
            "Authorization": f"Bearer {_drive_access_token()}",
            "Accept": accept,
        },
        method="GET",
    )
    with urllib.request.urlopen(request, timeout=30) as response:
        body = response.read()
        return body, dict(response.headers.items())


def _drive_json(path: str, query: dict[str, Any] | None = None) -> dict[str, Any]:
    body, _ = _drive_request(path, query=query)
    return json.loads(body.decode("utf-8"))


def _drive_text(file_id: str, mime_type: str) -> str:
    if mime_type.startswith("application/vnd.google-apps."):
        export_mime = {
            "application/vnd.google-apps.document": "text/plain",
            "application/vnd.google-apps.presentation": "text/plain",
            "application/vnd.google-apps.spreadsheet": "text/csv",
        }.get(mime_type)
        if not export_mime:
            raise ValueError(f"Unsupported Google Workspace export type: {mime_type}")
        body, _ = _drive_request(
            f"/drive/v3/files/{file_id}/export",
            query={"mimeType": export_mime},
            accept="text/plain",
        )
        return body.decode("utf-8", errors="ignore")

    body, _ = _drive_request(
        f"/drive/v3/files/{file_id}",
        query={"alt": "media"},
        accept="application/octet-stream",
    )
    return body.decode("utf-8", errors="ignore")


def _local_result_id(path: Path) -> str:
    return f"local:{_relative_path(path)}"


def _drive_result_id(file_id: str) -> str:
    return f"drive:{file_id}"


@mcp.tool()
def healthcheck() -> dict[str, Any]:
    """Return the current document root and file types this server can inspect."""
    return {
        "server": APP_NAME,
        "doc_root": str(_doc_root()),
        "allowed_extensions": sorted(_allowed_extensions()),
    }


@mcp.tool()
def list_documents(subdir: str = "", limit: int = 50) -> list[FileMatch]:
    """List documents under DOC_ROOT so the agent can discover what exists."""
    matches = _iter_document_paths(subdir=subdir)[: max(limit, 0)]
    return [
        FileMatch(
            path=_relative_path(path),
            name=path.name,
            extension=path.suffix.lower(),
            size_bytes=path.stat().st_size,
        )
        for path in matches
    ]


@mcp.tool()
def inspect_document(path: str) -> dict[str, Any]:
    """Return basic metadata and a short preview for one document."""
    resolved = _resolve_user_path(path)
    if not resolved.is_file():
        raise FileNotFoundError(f"Not a file: {path}")

    text = _extract_text(resolved)
    preview = " ".join(text.split())[:400]
    return {
        "path": _relative_path(resolved),
        "extension": resolved.suffix.lower(),
        "size_bytes": resolved.stat().st_size,
        "preview": preview,
        "character_count": len(text),
        "line_count": len(text.splitlines()),
    }


@mcp.tool()
def extract_document_text(path: str, max_chars: int = 12000) -> dict[str, Any]:
    """Extract text from a supported document so the agent can reason over it."""
    resolved = _resolve_user_path(path)
    if not resolved.is_file():
        raise FileNotFoundError(f"Not a file: {path}")

    text = _extract_text(resolved)
    truncated = text[: max(max_chars, 0)]
    return {
        "path": _relative_path(resolved),
        "truncated": len(truncated) < len(text),
        "character_count": len(text),
        "text": truncated,
    }


@mcp.tool()
def search_documents(
    query: str,
    subdir: str = "",
    limit: int = 10,
    preview_chars: int = 220,
) -> list[SearchMatch]:
    """Search across supported documents and return match counts plus snippets."""
    if not query.strip():
        raise ValueError("Query must not be empty.")

    results: list[SearchMatch] = []
    for path in _iter_document_paths(subdir=subdir):
        try:
            text = _extract_text(path)
        except Exception:
            continue
        count = text.lower().count(query.lower())
        if count:
            results.append(
                SearchMatch(
                    path=_relative_path(path),
                    occurrences=count,
                    snippet=_snippet(text, query, max_chars=preview_chars),
                )
            )

    results.sort(key=lambda item: (-item.occurrences, item.path))
    return results[: max(limit, 0)]


@mcp.tool()
def summarize_collection(subdir: str = "", limit: int = 200) -> dict[str, Any]:
    """Give the agent a quick overview of the document collection."""
    paths = _iter_document_paths(subdir=subdir)[: max(limit, 0)]
    extensions = Counter(path.suffix.lower() for path in paths)
    total_bytes = sum(path.stat().st_size for path in paths)
    return {
        "doc_root": str(_doc_root()),
        "subdir": subdir or ".",
        "file_count": len(paths),
        "total_size_bytes": total_bytes,
        "extensions": dict(sorted(extensions.items())),
        "sample_files": [_relative_path(path) for path in paths[:10]],
    }


@mcp.tool()
def propose_rename(path: str) -> RenameSuggestion:
    """Suggest a cleaner filename based on the document's extracted text."""
    resolved = _resolve_user_path(path)
    if not resolved.is_file():
        raise FileNotFoundError(f"Not a file: {path}")

    text = _extract_text(resolved)
    suggested_name, basis = _suggest_name_for_document(resolved, text)
    return RenameSuggestion(
        current_path=_relative_path(resolved),
        suggested_name=suggested_name,
        suggested_path=str(Path(_relative_path(resolved)).with_name(suggested_name)),
        basis=basis,
    )


@mcp.tool()
def move_document(
    source_path: str,
    destination_dir: str,
    new_name: str = "",
    overwrite: bool = False,
    dry_run: bool = True,
) -> dict[str, Any]:
    """Move a document inside DOC_ROOT, with a dry-run mode by default."""
    source = _resolve_user_path(source_path)
    if not source.is_file():
        raise FileNotFoundError(f"Not a file: {source_path}")

    destination_directory = _resolve_user_path(destination_dir)
    target_name = new_name.strip() or source.name
    target_name = _safe_slug(target_name, fallback=source.stem)
    if source.suffix and not target_name.lower().endswith(source.suffix.lower()):
        target_name = f"{target_name}{source.suffix.lower()}"
    target = destination_directory / target_name
    target = target.resolve()
    try:
        target.relative_to(_doc_root())
    except ValueError as exc:
        raise ValueError("Destination is outside DOC_ROOT.") from exc

    same_target = source.resolve() == target
    if same_target:
        return {
            "source_path": _relative_path(source),
            "destination_path": _relative_path(target),
            "dry_run": dry_run,
            "moved": False,
            "note": "Source and destination are the same file.",
        }

    if target.exists() and not overwrite:
        raise FileExistsError(f"Destination already exists: {_relative_path(target)}")

    result = {
        "source_path": _relative_path(source),
        "destination_path": _relative_path(target),
        "dry_run": dry_run,
        "moved": False,
    }
    if dry_run:
        return result

    target.parent.mkdir(parents=True, exist_ok=True)
    shutil.move(str(source), str(target))
    result["moved"] = True
    return result


@mcp.tool()
def drive_status() -> dict[str, Any]:
    """Report whether the optional Google Drive bridge has been configured."""
    token_present = bool(os.environ.get("GOOGLE_DRIVE_ACCESS_TOKEN", "").strip())
    return {
        "drive_bridge_enabled": token_present,
        "expected_env_vars": ["GOOGLE_DRIVE_ACCESS_TOKEN"],
        "notes": "Uses Google Drive API v3 with OAuth access tokens.",
    }


@mcp.tool()
def drive_search_files(
    query: str = "",
    page_size: int = 10,
    include_trashed: bool = False,
    shared_drive_id: str = "",
) -> dict[str, Any]:
    """Search Google Drive files through the Drive API when a token is configured."""
    q_parts: list[str] = []
    if query.strip():
        q_parts.append(query.strip())
    if not include_trashed:
        q_parts.append("trashed = false")

    request_query: dict[str, Any] = {
        "pageSize": max(1, min(page_size, 100)),
        "fields": "files(id,name,mimeType,modifiedTime,parents,driveId,size,webViewLink),nextPageToken",
        "q": " and ".join(q_parts) if q_parts else None,
        "includeItemsFromAllDrives": "true",
        "supportsAllDrives": "true",
        "orderBy": "modifiedTime desc",
    }
    if shared_drive_id.strip():
        request_query.update(
            {
                "corpora": "drive",
                "driveId": shared_drive_id.strip(),
            }
        )
    else:
        request_query["corpora"] = "user"

    return _drive_json("/drive/v3/files", query=request_query)


@mcp.tool()
def drive_get_file(file_id: str) -> dict[str, Any]:
    """Fetch metadata for one Drive file."""
    return _drive_json(
        f"/drive/v3/files/{file_id}",
        query={
            "fields": "id,name,mimeType,modifiedTime,parents,driveId,size,webViewLink,capabilities/canDownload",
            "supportsAllDrives": "true",
        },
    )


@mcp.tool()
def drive_extract_text(file_id: str, max_chars: int = 12000) -> dict[str, Any]:
    """Download or export Drive file text for Docs, Sheets, Slides, and text-like files."""
    metadata = drive_get_file(file_id)
    mime_type = metadata["mimeType"]
    text = _drive_text(file_id, mime_type)
    truncated = text[: max(max_chars, 0)]
    return {
        "file": metadata,
        "truncated": len(truncated) < len(text),
        "character_count": len(text),
        "text": truncated,
    }


@mcp.tool()
def search(query: str) -> str:
    """Use this when ChatGPT needs a search-style entry point over your documents."""
    results: list[dict[str, Any]] = []

    try:
        for match in search_documents(query=query, limit=8):
            local_path = _resolve_user_path(match.path)
            results.append(
                {
                    "id": _local_result_id(local_path),
                    "title": local_path.name,
                    "url": _local_document_url(local_path),
                }
            )
    except Exception:
        pass

    if os.environ.get("GOOGLE_DRIVE_ACCESS_TOKEN", "").strip():
        try:
            escaped_query = query.replace("'", "\\'")
            drive_results = drive_search_files(
                query=f"fullText contains '{escaped_query}'",
                page_size=5,
            )
            for item in drive_results.get("files", []):
                results.append(
                    {
                        "id": _drive_result_id(item["id"]),
                        "title": item["name"],
                        "url": item.get("webViewLink")
                        or f"https://drive.google.com/file/d/{item['id']}/view",
                    }
                )
        except Exception:
            pass

    return json.dumps({"results": results[:10]}, ensure_ascii=True)


@mcp.tool()
def fetch(id: str) -> str:
    """Use this when ChatGPT needs the full contents of one search result item."""
    if id.startswith("local:"):
        relative = id.removeprefix("local:")
        path = _resolve_user_path(relative)
        text = _extract_text(path)
        payload = {
            "id": id,
            "title": path.name,
            "text": text,
            "url": _local_document_url(path),
            "metadata": {
                "source": "local",
                "path": _relative_path(path),
                "extension": path.suffix.lower(),
            },
        }
        return json.dumps(payload, ensure_ascii=True)

    if id.startswith("drive:"):
        file_id = id.removeprefix("drive:")
        metadata = drive_get_file(file_id)
        payload = {
            "id": id,
            "title": metadata["name"],
            "text": _drive_text(file_id, metadata["mimeType"]),
            "url": metadata.get("webViewLink")
            or f"https://drive.google.com/file/d/{file_id}/view",
            "metadata": {
                "source": "google_drive",
                "mime_type": metadata["mimeType"],
                "modified_time": metadata.get("modifiedTime"),
            },
        }
        return json.dumps(payload, ensure_ascii=True)

    raise ValueError("Unknown result id. Expected local:<path> or drive:<file_id>.")


if __name__ == "__main__":
    host = os.environ.get("HOST", "0.0.0.0")
    port = int(os.environ.get("PORT", "8000"))
    mcp.settings.host = host
    mcp.settings.port = port
    mcp.run(transport="streamable-http")
